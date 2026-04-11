"""
Export training loss curves from training_history.json to an Excel workbook.

The workbook contains a data sheet with per-epoch metrics and a native Excel
line chart plotting the loss components over epochs.

Usage:
    python src/plot_training_history.py
    python src/plot_training_history.py --history trained_models/training_history.json
    python src/plot_training_history.py --output trained_models/loss_curves.xlsx
    python src/plot_training_history.py --log-scale
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

LOSS_KEYS = ["total", "attention", "mimicry", "relation", "segmentation"]
LOSS_LABELS = {
    "total": "Total",
    "attention": "Attention transfer",
    "mimicry": "Feature mimicry",
    "relation": "Relation (Gram)",
    "segmentation": "Segmentation (BCE+Dice)",
}


def load_history(path: Path) -> list[dict]:
    with open(path) as f:
        history = json.load(f)
    if not history:
        raise ValueError(f"{path} is empty — no epochs recorded yet.")
    history.sort(key=lambda e: e["epoch"])
    return history


def build_workbook(history: list[dict], log_scale: bool) -> Workbook:
    has_seg = any(e.get("segmentation", 0) > 0 for e in history)
    columns = ["epoch"] + [k for k in LOSS_KEYS if k != "segmentation" or has_seg] + ["lr"]

    wb = Workbook()
    ws = wb.active
    ws.title = "history"

    headers = ["Epoch"] + [LOSS_LABELS[k] for k in columns[1:-1]] + ["Learning rate"]
    ws.append(headers)
    for entry in history:
        ws.append([entry.get(k, 0.0) for k in columns])

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)

    n_rows = len(history)
    last_row = n_rows + 1

    chart = LineChart()
    chart.title = "Knowledge Distillation Training Loss"
    chart.x_axis.title = "Epoch"
    chart.y_axis.title = "Loss"
    chart.height = 11
    chart.width = 22
    chart.style = 2
    if log_scale:
        chart.y_axis.scaling.logBase = 10

    loss_col_start = 2
    loss_col_end = len(columns) - 1
    data = Reference(
        ws,
        min_col=loss_col_start,
        max_col=loss_col_end,
        min_row=1,
        max_row=last_row,
    )
    categories = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    for series in chart.series:
        series.smooth = False

    ws.add_chart(chart, f"{get_column_letter(len(columns) + 2)}2")

    lr_chart = LineChart()
    lr_chart.title = "Learning rate schedule"
    lr_chart.x_axis.title = "Epoch"
    lr_chart.y_axis.title = "LR"
    lr_chart.height = 7
    lr_chart.width = 22
    lr_chart.style = 2
    lr_col = len(columns)
    lr_data = Reference(ws, min_col=lr_col, max_col=lr_col, min_row=1, max_row=last_row)
    lr_chart.add_data(lr_data, titles_from_data=True)
    lr_chart.set_categories(categories)
    ws.add_chart(lr_chart, f"{get_column_letter(len(columns) + 2)}25")

    return wb


def main():
    parser = argparse.ArgumentParser(description="Export training loss curves to Excel.")
    parser.add_argument(
        "--history", type=Path,
        default=Path("trained_models/training_history.json"),
        help="Path to training_history.json",
    )
    parser.add_argument(
        "--output", type=Path,
        default=Path("trained_models/loss_curves.xlsx"),
        help="Output .xlsx path",
    )
    parser.add_argument("--log-scale", action="store_true", help="Use log scale on loss axis")
    args = parser.parse_args()

    if not args.history.exists():
        raise FileNotFoundError(f"History file not found: {args.history}")

    history = load_history(args.history)
    wb = build_workbook(history, args.log_scale)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    final = history[-1]
    has_seg = any(e.get("segmentation", 0) > 0 for e in history)
    print(f"Saved workbook to {args.output}")
    print(f"Latest epoch ({final['epoch']}): total={final.get('total', 0):.4f}"
          f"  att={final.get('attention', 0):.4f}"
          f"  mim={final.get('mimicry', 0):.4f}"
          f"  rel={final.get('relation', 0):.4f}"
          + (f"  seg={final.get('segmentation', 0):.4f}" if has_seg else ""))


if __name__ == "__main__":
    main()
